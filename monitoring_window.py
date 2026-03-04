from __future__ import annotations

import os
from typing import List, Tuple

from PyQt5.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QLineEdit,
    QTableWidget,
    QTableWidgetItem,
    QFileDialog,
    QMessageBox,
    QGroupBox,
    QDialog,
    QListWidget,
    QListWidgetItem,
    QDialogButtonBox,
    QHeaderView,
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon

from db import Database
from app_icon import get_icon_path


class _ClassSelectMenuHelper:
    """Вспомогательный объект для выбора класса через всплывающее меню."""

    def __init__(self, db: Database, parent: QWidget, line_edit: QLineEdit):
        self.db = db
        self.parent = parent
        self.line_edit = line_edit
        self._current_class_number: str = ""

    @property
    def current_class_number(self) -> str:
        return self._current_class_number

    def open_menu(self, button: QPushButton) -> None:
        from PyQt5.QtWidgets import QMenu

        forms = self.db.forms_get_all()
        if not forms:
            QMessageBox.information(
                self.parent,
                "Классы",
                "Сначала добавьте классы в разделе «Таблицы» → «Классы».",
            )
            return
        menu = QMenu(self.parent)
        for r in forms:
            number = r["number"]
            action = menu.addAction(number)
            action.triggered.connect(
                lambda checked=False, num=number: self._set_class(num)
            )
        menu.exec_(button.mapToGlobal(button.rect().bottomLeft()))

    def _set_class(self, number: str) -> None:
        self._current_class_number = number
        self.line_edit.setText(number)


class _PupilSelectDialog(QDialog):
    """Всплывающее окно выбора ученика по номеру класса."""

    def __init__(self, db: Database, class_number: str, parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Выбор ученика")
        layout = QVBoxLayout(self)

        self.list_widget = QListWidget()
        layout.addWidget(self.list_widget)

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)
        layout.addWidget(bb)

        # Находим form_id по номеру класса
        forms = {r["number"]: r["id"] for r in self.db.forms_get_all()}
        form_id = forms.get(class_number)
        if form_id is None:
            return

        rows = self.db.pupils_get_by_form_id(form_id)
        for r in rows:
            text = f"{r['surname']} {r['name']} {r['patronymic'] or ''}".strip()
            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, r)
            self.list_widget.addItem(item)

    def selected_pupil(self):
        item = self.list_widget.currentItem()
        if not item:
            return None
        return item.data(Qt.UserRole)


class _SpecialistSelectDialog(QDialog):
    """Выбор специалиста из таблицы experts."""

    def __init__(self, db: Database, current_name: str = "", parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Выбор специалиста")
        layout = QVBoxLayout(self)
        self.list_widget = QListWidget()
        layout.addWidget(self.list_widget)

        rows = self.db.experts_get_all()
        current_index = -1
        for i, r in enumerate(rows):
            name = r["name"] or ""
            item = QListWidgetItem(name)
            self.list_widget.addItem(item)
            if current_name and name == current_name:
                current_index = i
        if current_index >= 0:
            self.list_widget.setCurrentRow(current_index)

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)
        layout.addWidget(bb)

    def selected_name(self) -> str:
        item = self.list_widget.currentItem()
        return item.text().strip() if item else ""


class MonitoringWindow(QWidget):
    """Окно «Мониторинг» (п. 10.4 PROJECT.md)."""

    def __init__(self, db: Database, parent: QWidget | None = None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Мониторинг")

        icon_path = get_icon_path()
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        self._class_helper: _ClassSelectMenuHelper | None = None
        self._current_pupil = None

        layout = QVBoxLayout(self)

        # Верхний блок: выбор ученика и специалиста
        top_group = QGroupBox("Выбор ученика и специалиста")
        top_layout = QVBoxLayout(top_group)

        # Первый ряд: Класс, Найти, Специалист, Выбор
        row1 = QHBoxLayout()
        row1.addWidget(QLabel("Класс:"))
        self.class_edit = QLineEdit()
        self.class_edit.setReadOnly(True)
        self.class_edit.setPlaceholderText("выберите из списка")
        row1.addWidget(self.class_edit)
        self.btn_class = QPushButton("…")
        self.btn_class.setMaximumWidth(32)
        self.btn_class.setToolTip("Выбрать класс из справочника")
        row1.addWidget(self.btn_class)
        self._class_helper = _ClassSelectMenuHelper(self.db, self, self.class_edit)
        self.btn_class.clicked.connect(
            lambda: self._class_helper.open_menu(self.btn_class)
        )

        self.btn_find_pupil = QPushButton("Найти")
        self.btn_find_pupil.clicked.connect(self._on_find_pupil)
        row1.addWidget(self.btn_find_pupil)

        row1.addWidget(QLabel("Специалист:"))
        self.specialist_edit = QLineEdit()
        self.specialist_edit.setReadOnly(True)
        self.specialist_edit.setPlaceholderText("выберите специалиста")
        row1.addWidget(self.specialist_edit)
        self.btn_specialist = QPushButton("…")
        self.btn_specialist.setMaximumWidth(32)
        self.btn_specialist.setToolTip("Выбрать специалиста из справочника")
        self.btn_specialist.clicked.connect(self._on_choose_specialist)
        row1.addWidget(self.btn_specialist)

        self.btn_load_analysis = QPushButton("Выбор")
        self.btn_load_analysis.clicked.connect(self._on_load_analysis)
        row1.addWidget(self.btn_load_analysis)

        row1.addStretch()
        top_layout.addLayout(row1)

        # Второй ряд: ФИО, Выгрузить в Excel, Очистить
        row2 = QHBoxLayout()
        row2.addWidget(QLabel("Фамилия:"))
        self.surname_edit = QLineEdit()
        self.surname_edit.setReadOnly(True)
        row2.addWidget(self.surname_edit)
        row2.addWidget(QLabel("Имя:"))
        self.name_edit = QLineEdit()
        self.name_edit.setReadOnly(True)
        row2.addWidget(self.name_edit)
        row2.addWidget(QLabel("Отчество:"))
        self.patronymic_edit = QLineEdit()
        self.patronymic_edit.setReadOnly(True)
        row2.addWidget(self.patronymic_edit)

        self.btn_export = QPushButton("Выгрузить в Excel")
        self.btn_export.clicked.connect(self._on_export_excel)
        row2.addWidget(self.btn_export)

        self.btn_clear = QPushButton("Очистить")
        self.btn_clear.clicked.connect(self._on_clear_all)
        row2.addWidget(self.btn_clear)

        row2.addStretch()
        top_layout.addLayout(row2)

        layout.addWidget(top_group)

        # Временная таблица критериев и результатов
        self.table = QTableWidget()
        self.table.setColumnCount(1)
        self.table.setHorizontalHeaderLabels(["Критерий"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        layout.addWidget(self.table)

        self.setMinimumSize(900, 500)

    # --- Вспомогательные методы ---

    def _require_class_and_pupil(self) -> bool:
        if not (self._class_helper and self._class_helper.current_class_number):
            QMessageBox.warning(self, "Класс", "Сначала выберите класс.")
            return False
        if not self._current_pupil:
            QMessageBox.warning(self, "Ученик", "Сначала выберите ученика.")
            return False
        return True

    # --- Обработчики ---

    def _on_find_pupil(self) -> None:
        if not self._class_helper or not self._class_helper.current_class_number:
            QMessageBox.information(self, "Класс", "Сначала выберите класс.")
            return
        dlg = _PupilSelectDialog(
            self.db,
            self._class_helper.current_class_number,
            self,
        )
        if dlg.exec_() != QDialog.Accepted:
            return
        pupil = dlg.selected_pupil()
        if not pupil:
            return
        self._current_pupil = pupil
        self.surname_edit.setText(pupil["surname"] or "")
        self.name_edit.setText(pupil["name"] or "")
        self.patronymic_edit.setText(pupil["patronymic"] or "")

    def _on_choose_specialist(self) -> None:
        current = self.specialist_edit.text().strip()
        dlg = _SpecialistSelectDialog(self.db, current_name=current, parent=self)
        if dlg.exec_() == QDialog.Accepted:
            name = dlg.selected_name()
            if name:
                self.specialist_edit.setText(name)

    def _on_load_analysis(self) -> None:
        if not self._require_class_and_pupil():
            return
        specialist = self.specialist_edit.text().strip()
        if not specialist:
            QMessageBox.warning(self, "Специалист", "Выберите специалиста.")
            return

        class_number = self._class_helper.current_class_number if self._class_helper else ""
        surname = self.surname_edit.text().strip()
        name = self.name_edit.text().strip()
        patronymic = self.patronymic_edit.text().strip()

        result_cols, rows = self.db.analysis_get_results_for_pupil(
            class_number=class_number,
            surname=surname,
            name=name,
            patronymic=patronymic,
            specialist=specialist,
        )
        if not rows:
            QMessageBox.information(
                self,
                "Данные мониторинга",
                "Для выбранного ученика и специалиста нет записей в таблице анализа.",
            )
            self.table.setRowCount(0)
            self.table.setColumnCount(1)
            self.table.setHorizontalHeaderLabels(["Критерий"])
            return

        # Готовим заголовки: "Критерий" + колонки результатов
        headers = ["Критерий"] + [self._humanize_result_column(c) for c in result_cols]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            self.table.setItem(i, 0, QTableWidgetItem(r["criterion"] or ""))
            for j, col in enumerate(result_cols, start=1):
                val = r[col] if col in r.keys() else ""
                self.table.setItem(i, j, QTableWidgetItem(val or ""))

        self.table.resizeColumnsToContents()

    def _humanize_result_column(self, col: str) -> str:
        # result_I_2025_2026 -> "Результат I 2025-2026"
        if not col.startswith("result_"):
            return col
        rest = col[len("result_") :]  # e.g. "I_2025_2026"
        parts = rest.split("_")
        if len(parts) >= 3:
            period = parts[0]
            year1 = parts[1]
            year2 = parts[2]
            return f"Результат {period} {year1}-{year2}"
        return f"Результат {rest.replace('_', ' ')}"

    def _on_export_excel(self) -> None:
        if self.table.rowCount() == 0 or self.table.columnCount() <= 1:
            QMessageBox.information(
                self,
                "Выгрузка в Excel",
                "Нет данных для выгрузки. Сначала выполните выбор.",
            )
            return

        try:
            from openpyxl import Workbook
        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Не удалось загрузить библиотеку openpyxl для выгрузки в Excel:\n{e}",
            )
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить отчёт мониторинга",
            "monitoring.xlsx",
            "Файлы Excel (*.xlsx);;Все файлы (*)",
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Мониторинг"

        class_number = self._class_helper.current_class_number if self._class_helper else ""
        surname = self.surname_edit.text().strip()
        name = self.name_edit.text().strip()
        patronymic = self.patronymic_edit.text().strip()
        specialist = self.specialist_edit.text().strip()

        ws["A1"] = f"Класс: {class_number}"
        ws["A2"] = f"Ученик: {surname} {name} {patronymic}".strip()
        ws["A3"] = f"Специалист: {specialist}"

        # Заголовки таблицы
        headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
        start_row = 5
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=start_row, column=col_idx, value=header)

        # Данные
        for row_idx in range(self.table.rowCount()):
            for col_idx in range(self.table.columnCount()):
                item = self.table.item(row_idx, col_idx)
                value = item.text() if item else ""
                ws.cell(row=start_row + 1 + row_idx, column=col_idx + 1, value=value)

        try:
            wb.save(path)
            QMessageBox.information(
                self,
                "Выгрузка в Excel",
                f"Файл с результатами мониторинга сохранён:\n{path}",
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка сохранения",
                f"Не удалось сохранить файл Excel:\n{e}",
            )

    def _on_clear_all(self) -> None:
        """Очистить временную таблицу и все окошки."""
        self.table.setRowCount(0)
        self.table.setColumnCount(1)
        self.table.setHorizontalHeaderLabels(["Критерий"])

        self.class_edit.clear()
        self.specialist_edit.clear()
        self.surname_edit.clear()
        self.name_edit.clear()
        self.patronymic_edit.clear()

        self._current_pupil = None
        if self._class_helper:
            # сбрасываем выбранный номер класса
            self._class_helper._current_class_number = ""

